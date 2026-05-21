#!/usr/bin/env python3
"""
scan_sheet.py  –  OCR scanner pre papierové skóre-listy Kocky Sveta.

Načíta fotku papierového A4 záznamu, rozdelí ho na herné bloky podľa
layout_config.json, OCR-om prečíta kumulatívne skóre a uloží výsledok
do Excel súboru v rovnakom formáte, aký exportuje samotná aplikácia.

Použitie
--------
    python scan_sheet.py photo.jpg
    python scan_sheet.py photo.jpg --config layout_config.json
    python scan_sheet.py photo.jpg --output vysledok.xlsx
    python scan_sheet.py photo.jpg --debug          # uloží debug_*.png s mriežkou

Závislosti
----------
    pip install -r requirements.txt
    # + nainštaluj Tesseract:
    #   Linux : sudo apt install tesseract-ocr tesseract-ocr-slk
    #   macOS : brew install tesseract
    #   Win   : https://github.com/UB-Mannheim/tesseract/wiki
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

# ---------------------------------------------------------------------------
# Optional heavy imports – give clear errors if missing
# ---------------------------------------------------------------------------
try:
    import cv2
    import numpy as np
except ImportError:
    sys.exit("chýba opencv-python  →  pip install opencv-python-headless")

try:
    import pytesseract
    from PIL import Image
except ImportError:
    sys.exit("chýba pytesseract/Pillow  →  pip install pytesseract pillow")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("chýba openpyxl  →  pip install openpyxl")


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
DASH = "—"   # hodnota pre prázdne/vynechané kolo (ako v app exporte)


class CellValue:
    """Jedna bunka mriežky – číslo alebo DASH."""
    __slots__ = ("raw", "value")

    def __init__(self, raw: str):
        self.raw = raw.strip()
        cleaned = self.raw.replace(" ", "").replace(",", "").replace(".", "")
        # — alebo prázdno
        if cleaned in ("", "-", "–", "—", "---"):
            self.value = None
        else:
            try:
                self.value = int(cleaned)
            except ValueError:
                try:
                    self.value = int(float(cleaned))
                except ValueError:
                    self.value = None

    def excel_value(self):
        return DASH if self.value is None else self.value

    def __repr__(self):
        return f"CellValue({self.value!r})"


class GameData:
    def __init__(self, game_id: int, sheet: str, players: list[str]):
        self.id = game_id
        self.sheet = sheet
        self.players = players
        self.rounds: list[list[CellValue]] = []   # rounds[kolo][hráč]

    def winner(self, target: int) -> Optional[tuple[str, int]]:
        """Vráti (meno, skóre) víťaza – kto prvý dosiahol target."""
        n = len(self.players)
        for rnd in self.rounds:
            for pi in range(n):
                v = rnd[pi].value if pi < len(rnd) else None
                if v is not None and v >= target:
                    return self.players[pi], v
        # fallback – najvyššie skóre
        maxes = [
            max((r[i].value for r in self.rounds if i < len(r) and r[i].value is not None),
                default=0)
            for i in range(n)
        ]
        if any(maxes):
            best = max(range(n), key=lambda i: maxes[i])
            return self.players[best], maxes[best]
        return None

    def col_maxes(self) -> list[Optional[int]]:
        n = len(self.players)
        return [
            max((r[i].value for r in self.rounds if i < len(r) and r[i].value is not None),
                default=None)
            for i in range(n)
        ]


# ---------------------------------------------------------------------------
# Image utilities
# ---------------------------------------------------------------------------

def load_and_normalise(path: str, target_width: int = 2480) -> np.ndarray:
    """Načíta obrázok, zrovná orientáciu, voliteľne zmenší na target_width."""
    img = cv2.imread(path)
    if img is None:
        sys.exit(f"Obrázok sa nedá načítať: {path}")

    # EXIF auto-rotate
    try:
        from PIL import Image as PILImage
        pil = PILImage.open(path)
        exif = pil._getexif() or {}
        orientation = exif.get(274, 1)
        rotations = {3: 180, 6: 270, 8: 90}
        if orientation in rotations:
            img = cv2.rotate(img, {
                90:  cv2.ROTATE_90_COUNTERCLOCKWISE,
                180: cv2.ROTATE_180,
                270: cv2.ROTATE_90_CLOCKWISE,
            }[rotations[orientation]])
    except Exception:
        pass

    h, w = img.shape[:2]
    if w > target_width:
        scale = target_width / w
        img = cv2.resize(img, (target_width, int(h * scale)), interpolation=cv2.INTER_LANCZOS4)
    return img


def crop_zone(img: np.ndarray, zone: dict) -> np.ndarray:
    """Vyreže relatívnu zónu z obrázka."""
    H, W = img.shape[:2]
    x = int(zone["x"] * W)
    y = int(zone["y"] * H)
    w = int(zone["w"] * W)
    h = int(zone["h"] * H)
    return img[y:y+h, x:x+w].copy()


def preprocess_for_ocr(cell_img: np.ndarray) -> np.ndarray:
    """Grayscale + denoise + adaptívny prah pre lepší OCR čísel."""
    gray = cv2.cvtColor(cell_img, cv2.COLOR_BGR2GRAY)
    # Scale-up malé bunky (Tesseract radšej väčšie obrázky)
    h, w = gray.shape
    if w < 80:
        scale = max(2, 100 // w)
        gray = cv2.resize(gray, (w * scale, h * scale), interpolation=cv2.INTER_CUBIC)
    # Denoise
    gray = cv2.fastNlMeansDenoising(gray, h=15)
    # Adaptívny prah – funguje aj pri nerovnomernom osvetlení
    binary = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,
        blockSize=31, C=10
    )
    # Biely okraj – Tesseract to vyžaduje
    binary = cv2.copyMakeBorder(binary, 10, 10, 10, 10, cv2.BORDER_CONSTANT, value=255)
    return binary


# ---------------------------------------------------------------------------
# Grid detection
# ---------------------------------------------------------------------------

def detect_grid_lines(block: np.ndarray, n_cols: int, n_rows: int,
                      header_rows: int = 0) -> tuple[list[int], list[int]]:
    """
    Pokúsi sa detekovať mriežku cez HoughLines.
    Ak sa nedarí, padne späť na rovnomerné rozdelenie.

    Vracia (x_splits, y_splits) – súradnice ľavých/horných hrán každej bunky.
    Posledná položka = pravý/spodný okraj bloku.
    """
    gray = cv2.cvtColor(block, cv2.COLOR_BGR2GRAY)
    H, W = gray.shape

    # Vyostrenie + hrany
    sharp = cv2.GaussianBlur(gray, (0, 0), 1)
    sharp = cv2.addWeighted(gray, 1.5, sharp, -0.5, 0)
    edges = cv2.Canny(sharp, 30, 90)

    lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=int(W * 0.25),
                             minLineLength=int(W * 0.3), maxLineGap=int(W * 0.05))

    h_positions: list[int] = []
    v_positions: list[int] = []

    if lines is not None:
        for line in lines:
            x1, y1, x2, y2 = line[0]
            angle = abs(np.degrees(np.arctan2(y2 - y1, x2 - x1)))
            if angle < 10:          # horizontálna
                h_positions.append((y1 + y2) // 2)
            elif angle > 80:        # vertikálna
                v_positions.append((x1 + x2) // 2)

    def cluster(positions: list[int], expected: int, size: int) -> list[int]:
        if not positions:
            return list(range(0, size + 1, size // expected))
        positions = sorted(set(positions))
        gap = size / expected
        clusters: list[list[int]] = []
        cur: list[int] = [positions[0]]
        for p in positions[1:]:
            if p - cur[-1] < gap * 0.4:
                cur.append(p)
            else:
                clusters.append(cur)
                cur = [p]
        clusters.append(cur)
        midpoints = [int(np.mean(c)) for c in clusters]
        # doplň 0 a size ak chýbajú
        if midpoints[0] > gap * 0.3:
            midpoints.insert(0, 0)
        if midpoints[-1] < size - gap * 0.3:
            midpoints.append(size)
        # ak máme iný počet ako expected+1, padni na rovnomernú
        if len(midpoints) != expected + 1:
            return [int(i * size / expected) for i in range(expected + 1)]
        return midpoints

    data_rows = n_rows - header_rows
    y_splits = cluster(h_positions, n_rows, H)
    x_splits = cluster(v_positions, n_cols, W)

    return x_splits, y_splits


def even_splits(size: int, n: int) -> list[int]:
    return [int(i * size / n) for i in range(n + 1)]


# ---------------------------------------------------------------------------
# OCR
# ---------------------------------------------------------------------------

TESS_CONFIG = (
    "--psm 7 "                         # jedna riadka textu
    "--oem 1 "                         # LSTM engine
    "-c tessedit_char_whitelist=0123456789—-"
)


def ocr_cell(cell_img: np.ndarray) -> CellValue:
    """OCR jednej bunky – vráti CellValue."""
    proc = preprocess_for_ocr(cell_img)
    try:
        pil_img = Image.fromarray(proc)
        text = pytesseract.image_to_string(pil_img, config=TESS_CONFIG, lang="slk+eng")
    except Exception:
        try:
            text = pytesseract.image_to_string(proc, config=TESS_CONFIG)
        except Exception:
            text = ""
    return CellValue(text)


# ---------------------------------------------------------------------------
# Block extraction
# ---------------------------------------------------------------------------

def extract_game(img: np.ndarray, cfg: dict, debug: bool = False,
                 debug_prefix: str = "") -> GameData:
    """
    Vyreže blok z obrázka, rozdelí na mriežku a OCR-om prečíta hodnoty.
    Vracia GameData.
    """
    game_id    = cfg["id"]
    players    = cfg["players"]
    zone       = cfg["zone"]
    header_rows = cfg.get("header_rows", 1)   # koľko riadkov sú hlavičky (preskočia sa)

    block = crop_zone(img, zone)
    H, W  = block.shape[:2]

    # Odhadnutý počet riadkov: hlavičky + max 25 kôl
    n_cols = len(players)
    # Zistíme zo štruktúry – skúsime detekciu
    # Ak nemáme "rows" v config, odhadneme podľa výšky bunky ~30px
    n_rows_guess = cfg.get("rows", max(10, H // 28))

    x_splits, y_splits = detect_grid_lines(block, n_cols, n_rows_guess, header_rows)

    # Ak x_splits nedáva správny počet stĺpcov, prepočítaj
    if len(x_splits) - 1 != n_cols:
        x_splits = even_splits(W, n_cols)
    # Nechaj y_splits ako-je – počet riadkov môže byť premenlivý

    data_rows_y = y_splits[header_rows:]   # preskočíme hlavičkové riadky

    game = GameData(game_id, cfg["sheet"], players)
    debug_img = block.copy() if debug else None

    for ri in range(len(data_rows_y) - 1):
        y0, y1 = data_rows_y[ri], data_rows_y[ri + 1]
        row_vals: list[CellValue] = []
        all_none = True

        for ci in range(n_cols):
            x0, x1 = x_splits[ci], x_splits[ci + 1]
            cell = block[y0:y1, x0:x1]
            if cell.size == 0:
                row_vals.append(CellValue(""))
                continue
            val = ocr_cell(cell)
            row_vals.append(val)
            if val.value is not None:
                all_none = False

            if debug and debug_img is not None:
                cv2.rectangle(debug_img, (x0, y0), (x1, y1), (0, 200, 0), 1)
                label = str(val.value) if val.value is not None else "—"
                cv2.putText(debug_img, label, (x0 + 2, y0 + 16),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 255), 1)

        if not all_none:
            game.rounds.append(row_vals)

    if debug and debug_img is not None:
        out_path = f"{debug_prefix}debug_game{game_id}.png"
        cv2.imwrite(out_path, debug_img)
        print(f"  [debug] {out_path}")

    return game


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

GOLD  = "FFD700"
DARK  = "1A1A1A"
GREY  = "DDDDDD"


def _border():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_excel(games: list[GameData], target_score: int, out_path: str):
    """Zapíše všetky hry do Excel súboru v rovnakom formáte ako app export."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # odstrán default prázdny sheet

    # ── Prehľad ────────────────────────────────────────────────────────────
    ws_prehled = wb.create_sheet("Prehľad")
    headers = ["#", "Začiatok", "Koniec", "Cieľ", "Počet hráčov", "Hráči", "Víťaz", "Stav"]
    for ci, h in enumerate(headers, start=1):
        cell = ws_prehled.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color=DARK)
        cell.fill = PatternFill("solid", fgColor=GOLD)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    now_str = datetime.now().strftime("%-d. %B %Y o %H:%M")

    for game in games:
        winner_info = game.winner(target_score)
        winner_name = winner_info[0] if winner_info else ""
        row = [
            game.id,
            now_str,         # Začiatok – neznámy, použijeme teraz
            now_str,         # Koniec
            target_score,
            len(game.players),
            ", ".join(game.players),
            winner_name,
            "Dokončený" if winner_info else "Nedokončený",
        ]
        r = game.id + 1
        for ci, val in enumerate(row, start=1):
            cell = ws_prehled.cell(row=r, column=ci, value=val)
            cell.border = _border()
            if ci == 1:
                cell.alignment = Alignment(horizontal="center")

    # Auto-šírka stĺpcov Prehľad
    for col in ws_prehled.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_prehled.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    # ── Jednotlivé hry ─────────────────────────────────────────────────────
    for game in games:
        ws = wb.create_sheet(game.sheet)
        winner_info = game.winner(target_score)
        players = game.players
        n = len(players)

        # Riadok 1 – nadpis
        title = (
            f"Hra (naskenovaná) — {', '.join(players)}"
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 1)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=12, color=DARK)
        title_cell.fill = PatternFill("solid", fgColor=GOLD)
        title_cell.alignment = Alignment(horizontal="left")

        # Riadok 2 – prázdny (rovnako ako v app exporte)
        # Riadok 3 – hlavičky
        header_row = ["Kolo"] + players
        for ci, h in enumerate(header_row, start=1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=GREY)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()

        # Riadky 4+ – dáta
        for ri, rnd in enumerate(game.rounds, start=1):
            excel_row = 3 + ri
            ws.cell(row=excel_row, column=1, value=ri).border = _border()
            for ci, cv in enumerate(rnd, start=2):
                cell = ws.cell(row=excel_row, column=ci, value=cv.excel_value())
                cell.border = _border()
                if isinstance(cell.value, int):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="center")

        # Riadok MAX
        max_row = 3 + len(game.rounds) + 1
        ws.cell(row=max_row, column=1, value="MAX").font = Font(bold=True)
        ws.cell(row=max_row, column=1).border = _border()
        for ci, mx in enumerate(game.col_maxes(), start=2):
            cell = ws.cell(row=max_row, column=ci, value=mx)
            cell.font = Font(bold=True)
            cell.border = _border()

        # Riadok Víťaz
        if winner_info:
            wname, wscore = winner_info
            trophy_row = max_row + 1
            ws.merge_cells(start_row=trophy_row, start_column=1,
                           end_row=trophy_row, end_column=n + 1)
            trophy_cell = ws.cell(row=trophy_row, column=1,
                                  value=f"🏆 Víťaz: {wname} ({wscore})")
            trophy_cell.font = Font(bold=True, size=11)
            trophy_cell.fill = PatternFill("solid", fgColor=GOLD)
            trophy_cell.alignment = Alignment(horizontal="center")

        # Auto-šírka stĺpcov
        ws.column_dimensions["A"].width = 8
        for ci in range(2, n + 2):
            ws.column_dimensions[get_column_letter(ci)].width = 14

    wb.save(out_path)
    print(f"\n✓ Uložené do: {out_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="OCR scanner pre Kocky Sveta skóre-listy")
    ap.add_argument("image",          help="Cesta k fotke A4 papiera")
    ap.add_argument("--config",  "-c", default=str(Path(__file__).with_name("layout_config.json")),
                    help="JSON config s rozložením blokov (default: layout_config.json)")
    ap.add_argument("--output",  "-o", help="Výstupný .xlsx súbor (default: <meno_fotky>.xlsx)")
    ap.add_argument("--debug",   "-d", action="store_true",
                    help="Uloží debug_gameN.png s označenou mriežkou")
    ap.add_argument("--width",         type=int, default=2480,
                    help="Cieľová šírka obrázka pre normalizáciu (default 2480 px)")
    args = ap.parse_args()

    # Načítaj config
    config_path = Path(args.config)
    if not config_path.exists():
        sys.exit(f"Config nenájdený: {config_path}")
    with config_path.open(encoding="utf-8") as f:
        cfg = json.load(f)

    target_score = cfg.get("target_score", 10000)
    games_cfg    = cfg["games"]

    # Výstupný súbor
    img_path = Path(args.image)
    out_path = args.output or str(img_path.with_suffix(".xlsx"))

    # Nacítaj a normalizuj obrázok
    print(f"Načítavam: {args.image}")
    img = load_and_normalise(args.image, target_width=args.width)
    H, W = img.shape[:2]
    print(f"  rozmer po normalizácii: {W}×{H}")

    # Extrakcia hier
    games: list[GameData] = []
    for gcfg in games_cfg:
        print(f"\nSpracovávam {gcfg['sheet']} ({', '.join(gcfg['players'])}) …")
        game = extract_game(img, gcfg, debug=args.debug,
                            debug_prefix=str(img_path.parent) + "/")
        n_rounds = len(game.rounds)
        print(f"  → {n_rounds} kôl načítaných")
        for ri, rnd in enumerate(game.rounds, start=1):
            vals = [str(cv.value) if cv.value is not None else "—" for cv in rnd]
            print(f"    kolo {ri:2d}: {' | '.join(vals)}")
        games.append(game)

    if not games:
        sys.exit("Žiadne hry sa nepodarilo extrahovať.")

    # Zápis do Excelu
    print(f"\nZapisujem Excel …")
    write_excel(games, target_score, out_path)


if __name__ == "__main__":
    main()
